import glob
import json
import sqlite3
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable, Optional, Tuple

import pandas as pd
from openpyxl.utils.cell import get_column_letter

# dateparser опционально: если нет — просто не будет попыток распарсить published_raw
try:
    import dateparser
except Exception:
    dateparser = None


# ============================================================
# НАСТРОЙКИ (шапка)
# ============================================================
INPUT_GLOB = "aftershock_articles_*.sqlite"
OUT_DIR = "compact_out"

MAX_BYTES = 35 * 1024 * 1024  # 35 MB

# Что НЕ выводим в Excel (и в compact sqlite тоже не кладём)
# ВАЖНО: published_at мы всё равно читаем из исходной БД, но НЕ сохраняем в compact и НЕ выводим в Excel.
EXCLUDE_FROM_OUTPUT = {"sha1", "created_at", "published_at", "summary", "url"}

# Итоговые колонки compact sqlite (и Excel берёт подмножество из них)
OUT_COLUMNS = [
    "nid",
    "published_day",
    "direction",
    "direction_confidence",
    "title",
    "tags",
    "tags_json",  # в Excel не выводим, но в compact sqlite оставляем (может пригодиться)
]

# Сжатие текста
MAX_TITLE_CHARS = 180

# Экспорт
EXPORT_XLSX = True
EXPORT_CSV = False

ROWS_PER_XLSX_PART = 150_000

DATE_FORMAT = "%Y-%m-%d_%H-%M-%S"

# Авто-ширина колонок Excel (может замедлять на очень больших файлах)
AUTO_FIT_COLUMNS = True
AUTO_FIT_MAX_WIDTH = 50
# ============================================================


@dataclass
class RowOut:
    nid: int
    published_day: Optional[str]
    direction: Optional[str]
    direction_confidence: Optional[float]
    title: Optional[str]
    tags: Optional[str]
    tags_json: Optional[str]


def _safe_trunc(s: Optional[str], n: int) -> Optional[str]:
    if s is None:
        return None
    s = str(s).strip()
    if len(s) <= n:
        return s
    return s[:n].rstrip() + "..."


def _tags_to_text(tags_json: Optional[str]) -> Optional[str]:
    if not tags_json:
        return None
    try:
        arr = json.loads(tags_json)
        if isinstance(arr, list):
            return ", ".join(str(x) for x in arr if x)
    except Exception:
        return None
    return None


def _published_day_from_published_at(published_at: Optional[str]) -> Optional[str]:
    # published_at обычно ISO: '2025-12-31T19:18:00+03:00'
    if not published_at:
        return None
    s = str(published_at).strip()
    if len(s) >= 10 and s[4] == "-" and s[7] == "-":
        return s[:10]
    return None


def _published_day_fallback_from_raw(published_raw: Optional[str]) -> Optional[str]:
    # Опционально пытаемся парсить "сырую" дату, если published_at пуст
    if not published_raw:
        return None
    if dateparser is None:
        return None
    try:
        dt = dateparser.parse(
            str(published_raw),
            languages=["ru"],
            settings={"DATE_ORDER": "DMY", "RETURN_AS_TIMEZONE_AWARE": False},
        )
        if not dt:
            return None
        return dt.date().isoformat()
    except Exception:
        return None


def ensure_out_schema(db: sqlite3.Connection):
    db.execute("PRAGMA journal_mode=WAL;")
    db.execute("PRAGMA synchronous=NORMAL;")
    db.execute(
        """
        CREATE TABLE IF NOT EXISTS articles_compact (
            nid INTEGER PRIMARY KEY,
            published_day TEXT,
            direction TEXT,
            direction_confidence REAL,
            title TEXT,
            tags TEXT,
            tags_json TEXT
        );
        """
    )
    db.execute("CREATE INDEX IF NOT EXISTS idx_compact_published_day ON articles_compact(published_day);")
    db.execute("CREATE INDEX IF NOT EXISTS idx_compact_direction ON articles_compact(direction);")
    db.commit()


def open_out_db(path: Path) -> sqlite3.Connection:
    db = sqlite3.connect(str(path))
    ensure_out_schema(db)
    return db


def vacuum_db(db: sqlite3.Connection):
    db.commit()
    db.execute("VACUUM;")
    db.commit()


def out_db_name(part: int, date_str: str) -> str:
    return f"aftershock_compact_{part:03d}_{date_str}.sqlite"


def out_xlsx_name(part: int, date_str: str) -> str:
    return f"aftershock_compact_{part:03d}_{date_str}.xlsx"


def out_csv_name(part: int, date_str: str) -> str:
    return f"aftershock_compact_{part:03d}_{date_str}.csv"


def file_size(path: Path) -> int:
    return path.stat().st_size if path.exists() else 0


def _source_columns(con: sqlite3.Connection, table: str = "articles") -> set[str]:
    cur = con.cursor()
    cur.execute(f"PRAGMA table_info({table})")
    cols = {row[1] for row in cur.fetchall()}
    cur.close()
    return cols


def iter_source_rows(src_path: Path, batch: int = 2000) -> Iterable[dict]:
    """
    Читает нужные колонки из исходной таблицы articles.
    ВАЖНО: published_at читаем всегда (если есть), чтобы посчитать published_day,
    но в output/Excel published_at не попадёт.
    """
    con = sqlite3.connect(str(src_path))
    cur = con.cursor()

    cols = _source_columns(con, "articles")

    # минимально нужные поля
    # (published_raw — необязателен, используется как fallback)
    desired = ["nid", "published_at", "published_raw", "title", "tags_json", "direction", "direction_confidence"]

    select_exprs = []
    for c in desired:
        if c in cols:
            select_exprs.append(c)
        else:
            select_exprs.append(f"NULL AS {c}")

    cur.execute(f"SELECT {', '.join(select_exprs)} FROM articles ORDER BY nid")

    while True:
        rows = cur.fetchmany(batch)
        if not rows:
            break
        for r in rows:
            d = dict(zip(desired, r))
            yield d

    cur.close()
    con.close()


def transform_row(d: dict) -> RowOut:
    nid = int(d["nid"])
    title = _safe_trunc(d.get("title"), MAX_TITLE_CHARS)

    tags_json = d.get("tags_json")
    tags_txt = _tags_to_text(tags_json)

    published_at = d.get("published_at")
    published_raw = d.get("published_raw")

    published_day = _published_day_from_published_at(published_at)
    if not published_day:
        published_day = _published_day_fallback_from_raw(published_raw)

    return RowOut(
        nid=nid,
        published_day=published_day,
        direction=d.get("direction"),
        direction_confidence=d.get("direction_confidence"),
        title=title,
        tags=tags_txt,
        tags_json=tags_json,
    )


def insert_many(db: sqlite3.Connection, rows: list[RowOut]):
    db.executemany(
        """
        INSERT OR REPLACE INTO articles_compact
        (nid, published_day, direction, direction_confidence, title, tags, tags_json)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        [
            (r.nid, r.published_day, r.direction, r.direction_confidence, r.title, r.tags, r.tags_json)
            for r in rows
        ],
    )


def auto_adjust_columns(ws):
    for column in ws.columns:
        max_length = 0
        col_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, AUTO_FIT_MAX_WIDTH)


def export_one_compact_db_to_excel(compact_db_path: Path, xlsx_path: Path, max_rows_per_file: int):
    con = sqlite3.connect(str(compact_db_path))

    # В Excel выводим только нужное (без tags_json)
    cols_for_excel = [
        "nid",
        "published_day",
        "direction",
        "direction_confidence",
        "title",
        "tags",
    ]

    df_iter = pd.read_sql_query(
        f"SELECT {', '.join(cols_for_excel)} FROM articles_compact ORDER BY nid",
        con,
        chunksize=50_000,
    )

    written = 0
    with pd.ExcelWriter(str(xlsx_path), engine="openpyxl") as writer:
        startrow = 0
        for df in df_iter:
            if written >= max_rows_per_file:
                break

            remain = max_rows_per_file - written
            if len(df) > remain:
                df = df.iloc[:remain].copy()

            # Чтобы Excel “не терял” дату, сохраняем published_day как строку YYYY-MM-DD
            if "published_day" in df.columns:
                df["published_day"] = df["published_day"].astype(str).replace({"None": ""})

            df.to_excel(writer, sheet_name="articles", index=False, startrow=startrow, header=(startrow == 0))
            startrow += len(df) + (1 if startrow == 0 else 0)
            written += len(df)

    if AUTO_FIT_COLUMNS:
        from openpyxl import load_workbook
        wb = load_workbook(str(xlsx_path))
        ws = wb["articles"]
        auto_adjust_columns(ws)
        wb.save(str(xlsx_path))

    con.close()


def export_one_compact_db_to_csv(compact_db_path: Path, csv_path: Path):
    con = sqlite3.connect(str(compact_db_path))
    cols_for_csv = ["nid", "published_day", "direction", "direction_confidence", "title", "tags"]

    df_iter = pd.read_sql_query(
        f"SELECT {', '.join(cols_for_csv)} FROM articles_compact ORDER BY nid",
        con,
        chunksize=100_000,
    )
    first = True
    for df in df_iter:
        if "published_day" in df.columns:
            df["published_day"] = df["published_day"].astype(str).replace({"None": ""})
        df.to_csv(csv_path, index=False, encoding="utf-8", mode="w" if first else "a", header=first)
        first = False

    con.close()


def main():
    out_dir = Path(OUT_DIR)
    out_dir.mkdir(parents=True, exist_ok=True)

    run_stamp = datetime.now().strftime(DATE_FORMAT)
    print(f"Export started at: {run_stamp}")

    src_files = sorted(Path(p) for p in glob.glob(INPUT_GLOB))
    if not src_files:
        raise SystemExit(f"Не найдено файлов по маске: {INPUT_GLOB}")

    part = 1
    out_db_path = out_dir / out_db_name(part, run_stamp)
    out_db = open_out_db(out_db_path)

    buffer: list[RowOut] = []
    inserted_total = 0
    rotations = 0

    try:
        for src in src_files:
            print(f"Reading: {src.name}")
            for d in iter_source_rows(src, batch=2000):
                buffer.append(transform_row(d))

                if len(buffer) >= 2000:
                    insert_many(out_db, buffer)
                    out_db.commit()
                    inserted_total += len(buffer)
                    buffer.clear()

                    # проверяем размер, если вырос — пробуем сначала VACUUM (иногда сильно уменьшает)
                    if file_size(out_db_path) > MAX_BYTES:
                        vacuum_db(out_db)

                    if file_size(out_db_path) > MAX_BYTES:
                        print(f"Rotate: {out_db_path.name} > {MAX_BYTES} bytes -> next part")
                        out_db.close()
                        part += 1
                        rotations += 1
                        out_db_path = out_dir / out_db_name(part, run_stamp)
                        out_db = open_out_db(out_db_path)

        if buffer:
            insert_many(out_db, buffer)
            out_db.commit()
            inserted_total += len(buffer)
            buffer.clear()

        vacuum_db(out_db)

    finally:
        out_db.close()

    print(f"Done. Rows inserted: {inserted_total}. Parts: {part}. Rotations: {rotations}")

    for p in range(1, part + 1):
        compact_db_path = out_dir / out_db_name(p, run_stamp)
        if not compact_db_path.exists():
            continue

        if EXPORT_XLSX:
            xlsx_path = out_dir / out_xlsx_name(p, run_stamp)
            print(f"Exporting XLSX: {xlsx_path.name}")
            export_one_compact_db_to_excel(compact_db_path, xlsx_path, ROWS_PER_XLSX_PART)
            sz = file_size(xlsx_path)
            print(f"  ✓ XLSX size={sz / (1024 * 1024):.2f} MB")
            if sz > MAX_BYTES:
                print("  WARNING: XLSX > 35MB. Уменьшите ROWS_PER_XLSX_PART или MAX_TITLE_CHARS.")

        if EXPORT_CSV:
            csv_path = out_dir / out_csv_name(p, run_stamp)
            print(f"Exporting CSV: {csv_path.name}")
            export_one_compact_db_to_csv(compact_db_path, csv_path)
            sz = file_size(csv_path)
            print(f"  ✓ CSV size={sz / (1024 * 1024):.2f} MB")

    print(f"All saved to: {out_dir.absolute()}")


if __name__ == "__main__":
    main()
